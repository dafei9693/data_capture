import requests
import re
from openpyxl import Workbook
import pymysql


def getHtml(url):
    header = {
        ''
        'sec - fetch - mode': 'cors',
        'origin': 'https: // search.bilibili.com',
        'user - agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36 Edg/96.0.1054.57'
    }
    resp = requests.get(url,header)
    return resp.text

def getKeys(text):
    blocks = re.findall('video-list.*?</ul',text,re.S)
    lis = re.findall('li.*?</li',blocks[0],re.S)
    data = []
    for li in lis:
        row = []
        row.append(re.findall('title=".*?"',li,re.S)[0].replace('title="',''))
        row.append(re.findall('href=".*?"',li,re.S)[0].replace('href="//','').replace('"',''))
        row.append(re.findall('观看.*?<span',li,re.S)[0].replace('观看" class="so-icon watch-num"><i class="icon-playtime"></i>','').replace('><span','').replace('</span','').replace(' ','').replace("\n",''))
        row.append(re.findall('弹幕.*?<span',li,re.S)[0].replace('弹幕" class="so-icon hide"><i class="icon-subtitle"></i>','').replace('><span','').replace('</span','').replace(' ','').replace("\n",''))
        row.append(re.findall('上传时间.*?<span',li,re.S)[0].replace('上传时间" class="so-icon time"><i class="icon-date"></i>','').replace('><span','').replace('</span','').replace(' ','').replace("\n",''))
        data.append(row)
    return data

def getdata():
    with open('a.txt','rb') as f:
        text = f.read().decode('utf-8')

    keys = text.split()
    data=[]
    row = []
    for i in range(len(keys)):
        row.append(keys[i])
        if (i+1)%5 == 0:
            data.append(row)
            row = []

    return data


def getDatafromSql():
    conn = pymysql.connect(user='root', password='jyf20010121', database='sjsx')
    cur = conn.cursor()
    cur.execute("select * from comprehensive")
    data = cur.fetchall()
    conn.commit()
    data = list(data)
    for i in range(len(data)):
        if type(data[i]) == tuple:
            data[i] = list(data[i])
    return data


def save(data):
    conn = pymysql.connect( user='root', password='jyf20010121', database='sjsx')
    cur = conn.cursor()
    cur.execute("use sjsx")
    #cur.execute("create table comprehensive(title char(100),url char(100),playtime char(10),comment char(10),uploadtime char(15))")
    for i in range(len(data)):
        try:
            cur.execute("insert into comprehensive values('"+data[i][0]+"','"+data[i][1]+"','"+data[i][2]+"','"+data[i][3]+"','"+data[i][4]+"')")
        except:
            pass
    cur.close()
    conn.commit()

def saveAsXlsx():
    data = getdata()
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    for i in range(len(data)):
        for j in range(len(data[i])):
            ws.cell(i+1,j+1).value = data[i][j]

    wb.save('data.xlsx')






